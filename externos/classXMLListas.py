#
# LISTA DE DADOS CLIENTE E EMPRESA ARQUIVO EM MASSA


import openpyxl

from base.base import Base
from base.app import VarGerais, VarRede
from base.classAno import AnoResult
from base.classProcedimento import ProcResult

from cliente.classRede import Rede, RedeResult
from cliente.classCliente import CliResult
from cliente.classFiles import FilesResult


class XMLListas:

    def __init__(self, xml):
        self.xml = xml

    def __repr__(self):
        return self.xml

    lista_caminho = []
    lista_caminho_pc = []
    lista_arquivo_txt = []
    lista_arquivo_xml = []
    lista_sigla_empresa = []
    lista_num_empresa = []
    lista_sigla_cliente = []
    lista_num_cliente = []
    lista_arquivo_capa = []
    lista_capa_modelo = []

    def listar_cam_arq_em_massa(self):
        try:
            pasta_int_imp = VarGerais.pasta_interna_imp
            pasta_int_exp = VarGerais.pasta_interna_exp
            comex = RedeResult.tipo_comex
            pasta_interna = pasta_int_exp if comex == 'E' else pasta_int_imp
            caminho_comex = Rede.escolher_caminho_comex(
                self, RedeResult.tipo_comex)
            movto = RedeResult.tipo_movto
            ano = AnoResult.ano_completo
            caminho_movto = caminho_comex + movto + '\\' + ano + '\\'

            l_01, arquivo_capa_modelo = Base.pesquisar_existe_arquivo(
                Base.self, VarRede.caminho_modelo, VarRede.arquivo_em_massa)
            existe_arqcapa = l_01
            arq_massa = VarRede.arquivo_em_massa
            caminho_capa_modelo = VarRede.caminho_modelo + arq_massa
            diretorio = caminho_movto
            subdiretorio = '\\' + pasta_interna
            return existe_arqcapa, caminho_capa_modelo, diretorio, subdiretorio
        except AttributeError:
            Base.alertar_error_except(
                self, 'classXMLListas', 'listar_caminhos_arquivo_em_massa')

    def criar_listas_em_massa(
            self, caminho_capa_modelo, diretorio, subdiretorio):
        try:
            wb_obj = openpyxl.load_workbook(
                caminho_capa_modelo, data_only=True)
            sheet_obj = wb_obj['Planilha1']
            qtd = sheet_obj.max_row

            for i in range(2, qtd + 1):
                REFCLIENTE = sheet_obj.cell(row=i, column=1).value
                REFEMPRESA = sheet_obj.cell(row=i, column=2).value
                MODAL = str(sheet_obj.cell(row=i, column=3).value)
                NUMCLIENTE = str(sheet_obj.cell(row=i, column=4).value)
                SIGLACLIENTE = str(sheet_obj.cell(row=i, column=5).value)
                NUMEMPRESA = str(sheet_obj.cell(row=i, column=6).value)
                SIGLAEMPRESA = str(sheet_obj.cell(row=i, column=7).value)

                modal = VarRede.capa_base + MODAL
                arquivo_modelo = f'{modal} {VarGerais.apelido} - 0000.xlsx'
                caminho_modelo = arquivo_modelo
                XMLListas.lista_capa_modelo.append(caminho_modelo)

                arquivo_txt = f'{NUMCLIENTE} - DI.txt'
                XMLListas.lista_arquivo_txt.append(arquivo_txt)
                arquivo_xml = f'{NUMCLIENTE} - DI.xml'
                XMLListas.lista_arquivo_xml.append(arquivo_xml)
                XMLListas.lista_sigla_empresa.append(SIGLAEMPRESA)
                XMLListas.lista_num_empresa.append(NUMEMPRESA)
                XMLListas.lista_sigla_cliente.append(SIGLACLIENTE)
                XMLListas.lista_num_cliente.append(NUMCLIENTE)

                if REFCLIENTE is not None:
                    apelido = VarGerais.apelido
                    narq = apelido + ' - ' + SIGLACLIENTE + ' - ' + NUMCLIENTE
                    arquivo_capa = f'{modal} {narq}.xlsx'
                    XMLListas.lista_arquivo_capa.append(arquivo_capa)
                    caminho = f'{diretorio}{REFCLIENTE} - {REFEMPRESA}\\'
                    XMLListas.lista_caminho.append(caminho)
                    refemp = REFEMPRESA + subdiretorio
                    caminho_pc = f'{diretorio}{REFCLIENTE} - {refemp}\\'
                    XMLListas.lista_caminho_pc.append(caminho_pc)

            listas_caminho_em_massa = [
                                            XMLListas.lista_caminho,
                                            XMLListas.lista_caminho_pc,
                                            XMLListas.lista_arquivo_txt,
                                            XMLListas.lista_arquivo_xml,
                                            XMLListas.lista_sigla_empresa,
                                            XMLListas.lista_num_empresa,
                                            XMLListas.lista_sigla_cliente,
                                            XMLListas.lista_num_cliente,
                                            XMLListas.lista_arquivo_capa,
                                            XMLListas.lista_capa_modelo
                                        ]
            return listas_caminho_em_massa
        except AttributeError:
            Base.alertar_error_except(
                self, 'classXMLListas', 'criar_listas_em_massa')


class XMLListasResult:

    def __init__(self, xml):
        self.xml = xml

    def __repr__(self):
        return self.xml

    def definir_listas_caminho_em_massa(self):
        try:
            if CliResult.valida_cliente:
                l_01, arquivo_em_massa = Base.pesquisar_existe_arquivo(
                    Base.self, VarRede.caminho_modelo,
                    VarRede.arquivo_em_massa)
                existe_arquivo_em_massa = l_01

                if (existe_arquivo_em_massa
                        and (ProcResult.cod_proc == '12'
                             or ProcResult.cod_proc == '23')):
                    l_02, l_03, l_04, l_05 = XMLListas.listar_cam_arq_em_massa(
                        Base.self)
                    caminho_capa_modelo = l_03
                    # trocar o caminho dos processos
                    diretorio = l_04 + 'PROCESSOS' + '\\'
                    subdiretorio = l_05
                    listas_caminho_em_massa = XMLListas.criar_listas_em_massa(
                        Base.self, caminho_capa_modelo, diretorio,
                        subdiretorio)
                    lista_processos = []
                    for processo in listas_caminho_em_massa[7]:
                        if processo != 'None':
                            lista_processos.append(processo)
                    qtd_processos = len(lista_processos)

                elif (ProcResult.cod_proc == '11'
                      or ProcResult.cod_proc == '21'):

                    caminho_processo = FilesResult.processo + '\\'
                    caminho_processo_pc = FilesResult.processo_pc + '\\'
                    txt = f'{CliResult.dados_lista[0]} - DI.txt'
                    xml = f'{CliResult.dados_lista[0]} - DI.xml'
                    modal = CliResult.dados_lista[6]
                    sgl_cli = CliResult.dados_lista[8]
                    num_cliente = CliResult.dados_lista[0]
                    rede = VarRede.capa_base + modal + ' ' + VarGerais.apelido
                    arquivo_capa = f'{rede} - {sgl_cli} - {num_cliente}.xlsx'

                    lista_caminho = [caminho_processo]
                    lista_caminho_pc = [caminho_processo_pc]
                    lista_arquivo_txt = [txt]
                    lista_arquivo_xml = [xml]
                    lista_sigla_empresa = [CliResult.dados_lista[10]]
                    lista_num_empresa = [CliResult.dados_lista[1]]
                    lista_sigla_cliente = [sgl_cli]
                    lista_num_cliente = [num_cliente]
                    lista_arquivo_capa = [arquivo_capa]
                    lista_capa_modelo = [FilesResult.capa_modelo]
                    lista_processos = [num_cliente]
                    qtd_processos = len(lista_processos)

                    listas_caminho_em_massa = [
                                                    lista_caminho,
                                                    lista_caminho_pc,
                                                    lista_arquivo_txt,
                                                    lista_arquivo_xml,
                                                    lista_sigla_empresa,
                                                    lista_num_empresa,
                                                    lista_sigla_cliente,
                                                    lista_num_cliente,
                                                    lista_arquivo_capa,
                                                    lista_capa_modelo
                                                ]

                else:
                    listas_caminho_em_massa = [
                                                'Z',
                                                'Z',
                                                'Z',
                                                'Z',
                                                'Z',
                                                'Z',
                                                'Z',
                                                'Z',
                                                'Z',
                                                'Z'
                                                ]
                    lista_processos = []
                    qtd_processos = len(lista_processos)

            else:
                listas_caminho_em_massa = [
                                            'Z',
                                            'Z',
                                            'Z',
                                            'Z',
                                            'Z',
                                            'Z',
                                            'Z',
                                            'Z',
                                            'Z',
                                            'Z'
                                            ]
                lista_processos = []
                qtd_processos = len(lista_processos)

            return listas_caminho_em_massa, qtd_processos, lista_processos
        except AttributeError:
            Base.alertar_error_except(
                self, 'classXMLListas', 'definir_listas_caminho_em_massa')
