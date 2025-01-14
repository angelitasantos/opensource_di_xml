#
# RETIRAR AS INFORMAÇÕES DO ARQUIVO SAUDE (EXCEL)


import time
import pandas as pd
import openpyxl
import locale

from base.base import Base
from base.app import VarGerais, VarSaude

from cliente.classRede import Rede
from cliente.classCliente import CliResult
from cliente.classFiles import FilesResult


class Saude:

    def __init__(self, financeiro):
        self.financeiro = financeiro

    def __repr__(self):
        return self.financeiro

    if CliResult.valida_cliente:
        val_files = FilesResult.valida_files
        cli = CliResult.dados_lista[1]
        if Rede.proc_com_modal or CliResult.dados_lista[8] is None:
            processo = int(cli if CliResult.valida_cliente else 0)
            ref_empresa = '' if val_files is False else cli

    def extrair_dados(self, df, nome_banco):
        try:
            df.NF = df.NF.fillna('', inplace=False)
            df.Credito = df.Credito.fillna('', inplace=False)
            df.Debito = df.Debito.fillna('', inplace=False)

            df['Data'] = pd.to_datetime(df.Data)
            df = df.sort_values(by='Data')

            df_filtrado = df[df['Numero'] == str(Saude.processo)]
            new_df = df_filtrado.assign(Banco=nome_banco)
            return new_df
        except AttributeError:
            Base.alertar_error_except(self, 'classSaude', 'extrair_dados')

    def extrair_dados_saude(self):
        try:
            df_banco1 = pd.read_excel(
                                            VarSaude.caminho_completo,
                                            sheet_name=VarSaude.empresa_banco1,
                                            usecols=VarSaude.require_cols,
                                            skiprows=range(0, 1)
                                        )

            df_banco2 = pd.read_excel(
                                            VarSaude.caminho_completo,
                                            sheet_name=VarSaude.empresa_banco2,
                                            usecols=VarSaude.require_cols,
                                            skiprows=range(0, 1)
                                        )

            new_df_banco1 = Saude.extrair_dados(
                self, df_banco1, VarSaude.banco1)
            new_df_banco2 = Saude.extrair_dados(
                self, df_banco2, VarSaude.banco2)

            new_df = pd.merge(new_df_banco1, new_df_banco2, how='outer')
            new_df['Data'] = pd.to_datetime(new_df.Data)
            new_df = new_df.sort_values(by='Data')
            new_df['Data'] = new_df['Data'].dt.strftime('%d/%m/%Y')
            pd.set_option('float_format', '{:.2f}'.format)

            df_completo = []
            locale.setlocale(locale.LC_MONETARY, 'pt_BR.UTF-8')
            for index, rows in new_df.iterrows():
                rows.Credito = locale.currency(rows.Credito)
                rows.Debito = locale.currency(rows.Debito)
                rows.Credito = rows.Credito.replace('R$ ', '')
                rows.Debito = rows.Debito.replace('R$ ', '')
                my_list = [rows.Banco, rows.Data, rows.Credito, rows.Debito,
                           rows.Despesas, rows.AR, rows.Tipo, rows.Numero,
                           rows.NF]
                df_completo.append(my_list)
            return df_completo
        except AttributeError:
            Base.alertar_error_except(
                self, 'classSaude', 'extrair_dados_saude')

    def criar_aba_saude(self, wb_obj, arquivo_excel):
        try:
            wb_obj.create_sheet(index=1, title='SAUDE')
            sheet_obj = wb_obj['SAUDE']

            sheet_obj['A1'] = 'Banco'
            sheet_obj['B1'] = 'Data'
            sheet_obj['C1'] = 'Credito'
            sheet_obj['D1'] = 'Debito'
            sheet_obj['E1'] = 'Despesas'
            sheet_obj['F1'] = 'AR'
            sheet_obj['G1'] = 'Tipo'
            sheet_obj['H1'] = 'Numero'
            sheet_obj['I1'] = 'NF'

            data = Saude.extrair_dados_saude(self)
            for i in data:
                sheet_obj.append(i)

            wb_obj.save(arquivo_excel)
        except AttributeError:
            Base.alertar_error_except(self, 'classSaude', 'criar_aba_saude')

    def coletar_dados_filtrados(self):
        try:
            existe_arquivo_capa, arquivo_capa = Base.pesquisar_existe_arquivo(
                Base.self, FilesResult.processo_pc, FilesResult.capa_novo)
            caminho_capa = FilesResult.processo_pc + '\\' + arquivo_capa
            wb_obj = openpyxl.load_workbook(caminho_capa)
            sheetnames = wb_obj.sheetnames

            if 'SAUDE' in sheetnames:
                wb_obj.remove(wb_obj['SAUDE'])
                Saude.criar_aba_saude(self, wb_obj, caminho_capa)
            else:
                Saude.criar_aba_saude(self, wb_obj, caminho_capa)
        except AttributeError:
            Base.alertar_error_except(
                self, 'classSaude', 'coletar_dados_filtrados')

    def atualizar_planilha_saude(self):
        try:
            existe_arquivo_capa, arquivo_capa = Base.pesquisar_existe_arquivo(
                Base.self, FilesResult.processo_pc, FilesResult.capa_novo)

            if existe_arquivo_capa:
                while VarGerais.tentativas <= VarGerais.total_tentativas:
                    try:
                        Saude.coletar_dados_filtrados(self)
                        if VarGerais.tentativas == VarGerais.total_tentativas:
                            mensagem = 'PLANILHA ATUALIZADA !!!'
                            Base.alertar_pyautogui(self, mensagem)
                    except AttributeError:
                        total = VarGerais.total_tentativas
                        tentativas_validas = total - VarGerais.tentativas
                        mensagem = f'''PLANILHA ABERTA !!!\n
                        Você tem {tentativas_validas} tentativa(s)
                        para fechá-la!!!\nA planilha não foi atualizada !!!'''
                        Base.alertar_pyautogui(self, mensagem)
                    time.sleep(1)
                    VarGerais.tentativas += 1

            elif not existe_arquivo_capa:
                mensagem = 'Arquivos não estão salvos na pasta PC do processo!'
                Base.alertar_pyautogui(self, mensagem)
                time.sleep(Base.time_sleep_1)
        except AttributeError:
            Base.alertar_error_except(
                self, 'classSaude', 'atualizar_planilha_saude')
